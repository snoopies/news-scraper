#!/usr/bin/env python3
"""
News Scraper - 获取新闻标题及链接，翻译为中文
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import json
import re
from deep_translator import GoogleTranslator
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def translate_to_chinese(text):
    """翻译文本为中文"""
    try:
        # 限制翻译长度，避免超时
        if len(text) > 500:
            text = text[:500]
        
        translator = GoogleTranslator(source='auto', target='zh-CN')
        translated = translator.translate(text)
        return translated
    except Exception as e:
        print(f"Translation error: {e}")
        return text  # 翻译失败返回原文

def get_nyt_news():
    """获取纽约时报最新新闻 - 通过 RSS-like 页面"""
    # 使用 NYT 的国际版首页
    url = "https://www.nytimes.com/section/world"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate',
    }

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')

        news_items = []
        
        # NYT 页面结构
        articles = soup.find_all(['article', 'li'])
        
        for article in articles:
            # 尝试多种选择器
            title_elem = article.find(['h2', 'h3', 'h4', 'a'], class_=lambda x: x and any(k in str(x).lower() for k in ['headline', 'title', 'story', 'link']))
            
            if not title_elem:
                # 尝试找所有链接
                links = article.find_all('a', href=True)
                for link in links:
                    href = link.get('href', '')
                    text = link.get_text(strip=True)
                    # 过滤有效新闻链接和标题
                    if text and len(text) > 20 and ('/202' in href or '/world/' in href or '/us/' in href):
                        if href.startswith('/'):
                            href = 'https://www.nytimes.com' + href
                        news_items.append({
                            'source': 'New York Times',
                            'title': text[:200],  # 限制标题长度
                            'link': href,
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        })
                        break
            else:
                title = title_elem.get_text(strip=True)
                link = article.find('a', href=True)
                if link and title:
                    href = link.get('href', '')
                    if href.startswith('/'):
                        href = 'https://www.nytimes.com' + href
                    news_items.append({
                        'source': 'New York Times',
                        'title': title[:200],
                        'link': href,
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
        
        # 去重
        seen = set()
        unique_items = []
        for item in news_items:
            if item['link'] not in seen:
                seen.add(item['link'])
                unique_items.append(item)
        
        return unique_items[:20]

    except Exception as e:
        print(f"Error fetching NYT news: {e}")
        return []

def get_aol_news():
    """获取 AOL 新闻"""
    url = "https://www.aol.com/news/"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
    }

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')

        news_items = []
        
        # AOL 页面结构
        for link in soup.find_all('a', href=True):
            href = link.get('href', '')
            text = link.get_text(strip=True)
            
            # 过滤有效新闻链接
            if text and len(text) > 15:
                # AOL 新闻链接特征
                if any(pattern in href for pattern in ['news', 'article', 'story', '/20']):
                    if not any(skip in href for skip in ['javascript:', '#', 'mailto:', 'advert']):
                        # 确保是完整URL
                        if href.startswith('/'):
                            href = 'https://www.aol.com' + href
                        
                        news_items.append({
                            'source': 'AOL',
                            'title': text[:200],
                            'link': href,
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        })
        
        # 去重
        seen = set()
        unique_items = []
        for item in news_items:
            if item['link'] not in seen and len(item['title']) > 20:
                seen.add(item['link'])
                unique_items.append(item)
        
        return unique_items[:20]

    except Exception as e:
        print(f"Error fetching AOL news: {e}")
        return []

def get_reuters_news():
    """获取路透社新闻作为补充"""
    url = "https://www.reuters.com/world/"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    }

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')

        news_items = []
        
        for link in soup.find_all('a', href=True):
            href = link.get('href', '')
            text = link.get_text(strip=True)
            
            if text and len(text) > 20 and '/world/' in href:
                if href.startswith('/'):
                    href = 'https://www.reuters.com' + href
                
                news_items.append({
                    'source': 'Reuters',
                    'title': text[:200],
                    'link': href,
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
        
        seen = set()
        unique_items = []
        for item in news_items:
            if item['link'] not in seen:
                seen.add(item['link'])
                unique_items.append(item)
        
        return unique_items[:15]

    except Exception as e:
        print(f"Error fetching Reuters news: {e}")
        return []

def get_bbc_news():
    """获取 BBC 新闻作为补充"""
    url = "https://www.bbc.com/news"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    }

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')

        news_items = []
        
        for link in soup.find_all('a', href=True):
            href = link.get('href', '')
            text = link.get_text(strip=True)
            
            if text and len(text) > 20:
                if '/news/' in href and not any(skip in href for skip in ['have-your-say', 'av/', 'live']):
                    if href.startswith('/'):
                        href = 'https://www.bbc.com' + href
                    
                    news_items.append({
                        'source': 'BBC',
                        'title': text[:200],
                        'link': href,
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
        
        seen = set()
        unique_items = []
        for item in news_items:
            if item['link'] not in seen:
                seen.add(item['link'])
                unique_items.append(item)
        
        return unique_items[:15]

    except Exception as e:
        print(f"Error fetching BBC news: {e}")
        return []

def beautify_excel(file_path):
    """美化Excel表格"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 定义样式
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        body_font = Font(name='Arial', size=10)
        body_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15  # source
        ws.column_dimensions['B'].width = 50  # title
        ws.column_dimensions['C'].width = 50  # title_zh
        ws.column_dimensions['D'].width = 60  # link
        ws.column_dimensions['E'].width = 20  # timestamp
        
        # 美化表头
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 美化数据行
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = body_alignment
                cell.border = thin_border
                
                # 为不同来源设置不同的背景色
                source_cell = row[0]  # source列
                if source_cell.value == 'New York Times':
                    row[0].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                elif source_cell.value == 'AOL':
                    row[0].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                elif source_cell.value == 'BBC':
                    row[0].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                elif source_cell.value == 'Reuters':
                    row[0].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存
        wb.save(file_path)
        print(f"✅ Excel表格已美化")
        
    except Exception as e:
        print(f"⚠️ 美化Excel时出错: {e}")

def main():
    print("=" * 60)
    print("🌍 新闻抓取器 - 支持中文翻译")
    print("=" * 60)

    # 获取新闻
    print("\n📰 正在获取纽约时报新闻...")
    nyt_news = get_nyt_news()
    print(f"   获取到 {len(nyt_news)} 条")
    
    print("\n📰 正在获取 AOL 新闻...")
    aol_news = get_aol_news()
    print(f"   获取到 {len(aol_news)} 条")
    
    print("\n📰 正在获取路透社新闻...")
    reuters_news = get_reuters_news()
    print(f"   获取到 {len(reuters_news)} 条")
    
    print("\n📰 正在获取 BBC 新闻...")
    bbc_news = get_bbc_news()
    print(f"   获取到 {len(bbc_news)} 条")

    # 合并新闻
    all_news = nyt_news + aol_news + reuters_news + bbc_news

    print(f"\n{'=' * 60}")
    print(f"📊 共获取 {len(all_news)} 条新闻")
    print(f"{'=' * 60}")

    # 创建 DataFrame
    if all_news:
        print(f"\n🔄 正在翻译标题为中文...")
        
        # 翻译标题
        for i, item in enumerate(all_news):
            print(f"   翻译中... {i+1}/{len(all_news)}", end='\r')
            item['title_zh'] = translate_to_chinese(item['title'])
        
        print(f"\n   ✅ 翻译完成")
        
        df = pd.DataFrame(all_news)

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = f'news_list_{timestamp}.xlsx'

        # 保存到 Excel
        df.to_excel(excel_file, index=False, engine='openpyxl')
        print(f"\n✅ Excel 文件已生成: {excel_file}")
        
        # 美化Excel
        print(f"\n🎨 正在美化Excel表格...")
        beautify_excel(excel_file)
        
        # 输出摘要
        print(f"\n📊 新闻统计:")
        source_counts = df['source'].value_counts()
        for source, count in source_counts.items():
            print(f"   {source}: {count} 条")

    else:
        # 即使没有新闻也生成空文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = f'news_list_{timestamp}.xlsx'
        df = pd.DataFrame(columns=['source', 'title', 'title_zh', 'link', 'timestamp'])
        df.to_excel(excel_file, index=False, engine='openpyxl')
        print(f"\n⚠️ 未获取到新闻，已生成空文件: {excel_file}")

    # 输出 JSON 格式（用于 GitHub Actions）
    output = {
        'total': len(all_news),
        'nyt_count': len(nyt_news),
        'aol_count': len(aol_news),
        'reuters_count': len(reuters_news),
        'bbc_count': len(bbc_news),
        'excel_file': excel_file,
        'news': all_news
    }

    with open('news_output.json', 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n🎉 新闻抓取完成！")
    print(f"📁 文件: {excel_file}")

if __name__ == '__main__':
    main()